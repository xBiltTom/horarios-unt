import io
import zipfile
from datetime import time

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from auth import get_current_active_user, require_secretaria
from database import get_db
from models import (
    AsignacionCarga, Aula, BloqueHorario, CargaNoLectiva, Curso,
    Docente, Laboratorio, RolEnum, Semestre, TurnoLaboratorio, Usuario,
)

router = APIRouter(prefix="/api/documentos", tags=["documentos"])

DIAS = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado"]
DIAS_LABEL = {"lunes": "LUN", "martes": "MAR", "miercoles": "MIE",
               "jueves": "JUE", "viernes": "VIE", "sabado": "SAB"}
TIPO_LABEL = {
    "teoria": "T", "practica": "P", "laboratorio": "L",
    "preparacion": "Prep", "consejeria": "Cons",
    "investigacion": "Inv", "rsu": "RSU",
    "asesoria": "Ases", "capacitacion": "Cap",
}
TIPO_COLOR = {
    "teoria": (0.87, 0.92, 1.0),
    "practica": (0.87, 1.0, 0.87),
    "laboratorio": (1.0, 0.95, 0.8),
    "preparacion": (0.95, 0.87, 1.0),
    "consejeria": (1.0, 0.87, 0.87),
    "investigacion": (0.87, 0.97, 0.97),
    "rsu": (1.0, 1.0, 0.8),
    "asesoria": (0.97, 0.92, 0.87),
    "capacitacion": (0.9, 0.9, 0.9),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_semestre(db: Session, semestre_id: int) -> Semestre:
    s = db.get(Semestre, semestre_id)
    if not s:
        raise HTTPException(status_code=404, detail="Semestre no encontrado")
    return s


def _get_docente(db: Session, docente_id: int) -> Docente:
    d = (
        db.query(Docente)
        .options(joinedload(Docente.departamento))
        .filter(Docente.id == docente_id)
        .first()
    )
    if not d:
        raise HTTPException(status_code=404, detail="Docente no encontrado")
    return d


def _time_to_float(t: time) -> float:
    return t.hour + t.minute / 60


def _bloques_docente(db: Session, docente_id: int, semestre_id: int) -> list[BloqueHorario]:
    return (
        db.query(BloqueHorario)
        .options(
            joinedload(BloqueHorario.aula),
            joinedload(BloqueHorario.laboratorio),
            joinedload(BloqueHorario.asignacion).joinedload(AsignacionCarga.curso),
        )
        .filter(
            BloqueHorario.docente_id == docente_id,
            BloqueHorario.semestre_id == semestre_id,
        )
        .order_by(BloqueHorario.dia, BloqueHorario.hora_inicio)
        .all()
    )


def _asignaciones_docente(db: Session, docente_id: int, semestre_id: int) -> list[AsignacionCarga]:
    return (
        db.query(AsignacionCarga)
        .options(
            joinedload(AsignacionCarga.curso),
            joinedload(AsignacionCarga.turnos_laboratorio),
        )
        .filter(
            AsignacionCarga.docente_id == docente_id,
            AsignacionCarga.semestre_id == semestre_id,
        )
        .all()
    )


def _carga_no_lectiva(db: Session, docente_id: int, semestre_id: int) -> list[CargaNoLectiva]:
    return (
        db.query(CargaNoLectiva)
        .filter(
            CargaNoLectiva.docente_id == docente_id,
            CargaNoLectiva.semestre_id == semestre_id,
            CargaNoLectiva.horas_asignadas > 0,
        )
        .all()
    )


# ---------------------------------------------------------------------------
# PDF generation helpers
# ---------------------------------------------------------------------------

def _build_horario_pdf(semestre: Semestre, ciclo: int, asignaciones: list, bloques: list) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Normal"],
                                 fontSize=13, alignment=TA_CENTER, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=10, alignment=TA_CENTER)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7)

    story = []
    story.append(Paragraph("UNIVERSIDAD NACIONAL DE TUCUMÁN", title_style))
    story.append(Paragraph(f"Horario de Clases — Ciclo {ciclo} — Semestre {semestre.numero} {semestre.anio}", sub_style))
    story.append(Spacer(1, 0.3*cm))

    # Course table
    if asignaciones:
        ctable_data = [["Curso", "Código", "T", "P", "Lab", "Docente"]]
        for asig in asignaciones:
            c = asig.curso
            d = asig.docente
            ctable_data.append([
                Paragraph(c.nombre, cell_style),
                c.codigo,
                str(c.horas_teoria * asig.grupos_teoria) if asig.dicta_teoria else "—",
                str(c.horas_practica * asig.grupos_practica) if asig.dicta_practica else "—",
                str(c.horas_laboratorio) if c.horas_laboratorio else "—",
                Paragraph(f"{d.apellidos}, {d.nombre}", cell_style),
            ])
        ctable = Table(ctable_data, colWidths=[6*cm, 2.5*cm, 1*cm, 1*cm, 1*cm, 5*cm])
        ctable.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A6CF7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(ctable)
        story.append(Spacer(1, 0.4*cm))

    # Weekly grid
    HORAS = list(range(7, 21))
    header = ["Hora"] + [DIAS_LABEL[d] for d in DIAS]
    grid_data = [header]
    for h in HORAS:
        row = [f"{h:02d}:00"]
        for dia in DIAS:
            blocks_here = [
                b for b in bloques
                if b.dia == dia and _time_to_float(b.hora_inicio) <= h < _time_to_float(b.hora_fin)
            ]
            if blocks_here:
                b = blocks_here[0]
                label = TIPO_LABEL.get(b.tipo, b.tipo)
                if b.asignacion:
                    label += f"\n{b.asignacion.curso.codigo}"
                if b.aula:
                    label += f"\n{b.aula.nombre}"
                elif b.laboratorio:
                    label += f"\n{b.laboratorio.nombre}"
                row.append(Paragraph(label, cell_style))
            else:
                row.append("")
        grid_data.append(row)

    col_w = [1.5*cm] + [3.5*cm] * 6
    grid = Table(grid_data, colWidths=col_w, rowHeights=[0.55*cm] * (len(HORAS) + 1))
    ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F9FAFB")),
    ])
    for ri, row in enumerate(grid_data[1:], 1):
        for ci, cell in enumerate(row[1:], 1):
            if cell:
                # Find the bloque to get its tipo for color
                dia = DIAS[ci - 1]
                h = HORAS[ri - 1]
                b_list = [
                    b for b in bloques
                    if b.dia == dia and _time_to_float(b.hora_inicio) <= h < _time_to_float(b.hora_fin)
                ]
                if b_list:
                    rgb = TIPO_COLOR.get(b_list[0].tipo, (1, 1, 1))
                    ts.add("BACKGROUND", (ci, ri), (ci, ri), colors.Color(*rgb))
    grid.setStyle(ts)
    story.append(grid)

    doc.build(story)
    return buf.getvalue()


def _build_horario_excel_sheet(ws, semestre, ciclo, asignaciones, bloques):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Horario — Ciclo {ciclo} — Semestre {semestre.numero} {semestre.anio}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Course table header
    row = 3
    headers = ["Curso", "Código", "T", "P", "Lab", "Docente"]
    header_fill = PatternFill("solid", fgColor="4A6CF7")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row += 1
    for asig in asignaciones:
        c = asig.curso
        d = asig.docente
        values = [
            c.nombre,
            c.codigo,
            c.horas_teoria * asig.grupos_teoria if asig.dicta_teoria else None,
            c.horas_practica * asig.grupos_practica if asig.dicta_practica else None,
            c.horas_laboratorio if c.horas_laboratorio else None,
            f"{d.apellidos}, {d.nombre}",
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
        row += 1

    # Grid header
    row += 1
    grid_header_fill = PatternFill("solid", fgColor="374151")
    ws.cell(row=row, column=1, value="Hora").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = grid_header_fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    for ci, dia in enumerate(DIAS, 2):
        cell = ws.cell(row=row, column=ci, value=DIAS_LABEL[dia])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = grid_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    TIPO_FILL = {k: PatternFill("solid", fgColor=
                  "%02X%02X%02X" % (int(v[0]*255), int(v[1]*255), int(v[2]*255)))
                 for k, v in TIPO_COLOR.items()}

    row += 1
    for h in range(7, 21):
        ws.cell(row=row, column=1, value=f"{h:02d}:00").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = border
        for ci, dia in enumerate(DIAS, 2):
            blocks_here = [
                b for b in bloques
                if b.dia == dia and _time_to_float(b.hora_inicio) <= h < _time_to_float(b.hora_fin)
            ]
            cell = ws.cell(row=row, column=ci)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if blocks_here:
                b = blocks_here[0]
                label = TIPO_LABEL.get(b.tipo, b.tipo)
                if b.asignacion:
                    label += f"\n{b.asignacion.curso.codigo}"
                cell.value = label
                cell.fill = TIPO_FILL.get(b.tipo, PatternFill())
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 10
    for ci in range(2, 8):
        ws.column_dimensions[get_column_letter(ci)]
    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/horario/{semestre_id}/pdf")
def horario_pdf(
    semestre_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_secretaria),
):
    semestre = _get_semestre(db, semestre_id)
    asignaciones_all = (
        db.query(AsignacionCarga)
        .options(
            joinedload(AsignacionCarga.curso),
            joinedload(AsignacionCarga.docente),
        )
        .filter(AsignacionCarga.semestre_id == semestre_id)
        .all()
    )
    bloques_all = (
        db.query(BloqueHorario)
        .options(
            joinedload(BloqueHorario.aula),
            joinedload(BloqueHorario.laboratorio),
            joinedload(BloqueHorario.asignacion).joinedload(AsignacionCarga.curso),
        )
        .filter(BloqueHorario.semestre_id == semestre_id)
        .all()
    )

    ciclos = sorted({a.curso.ciclo for a in asignaciones_all})
    if not ciclos:
        ciclos = [0]

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for ciclo in ciclos:
            asig_ciclo = [a for a in asignaciones_all if a.curso.ciclo == ciclo]
            bloques_ciclo = [b for b in bloques_all if b.asignacion and b.asignacion.curso.ciclo == ciclo]
            pdf_bytes = _build_horario_pdf(semestre, ciclo, asig_ciclo, bloques_ciclo)
            zf.writestr(f"horario_ciclo_{ciclo}.pdf", pdf_bytes)

    zip_buf.seek(0)
    fname = f"horarios_{semestre.numero}_{semestre.anio}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/horario/{semestre_id}/excel")
def horario_excel(
    semestre_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_secretaria),
):
    import openpyxl

    semestre = _get_semestre(db, semestre_id)
    asignaciones_all = (
        db.query(AsignacionCarga)
        .options(
            joinedload(AsignacionCarga.curso),
            joinedload(AsignacionCarga.docente),
        )
        .filter(AsignacionCarga.semestre_id == semestre_id)
        .all()
    )
    bloques_all = (
        db.query(BloqueHorario)
        .options(
            joinedload(BloqueHorario.aula),
            joinedload(BloqueHorario.laboratorio),
            joinedload(BloqueHorario.asignacion).joinedload(AsignacionCarga.curso),
        )
        .filter(BloqueHorario.semestre_id == semestre_id)
        .all()
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ciclos = sorted({a.curso.ciclo for a in asignaciones_all})
    if not ciclos:
        ciclos = [0]

    for ciclo in ciclos:
        ws = wb.create_sheet(title=f"Ciclo {ciclo}")
        asig_ciclo = [a for a in asignaciones_all if a.curso.ciclo == ciclo]
        bloques_ciclo = [b for b in bloques_all if b.asignacion and b.asignacion.curso.ciclo == ciclo]
        _build_horario_excel_sheet(ws, semestre, ciclo, asig_ciclo, bloques_ciclo)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"horarios_{semestre.numero}_{semestre.anio}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )




# ---------------------------------------------------------------------------
# Auth helper — docente can only access own documents
# ---------------------------------------------------------------------------

def _check_docente_or_staff(current: Usuario, docente_id: int, db: Session):
    if current.rol == RolEnum.docente:
        own = db.query(Docente).filter(Docente.usuario_id == current.id).first()
        if not own or own.id != docente_id:
            raise HTTPException(status_code=403, detail="No tiene permisos")
    elif current.rol not in (RolEnum.admin, RolEnum.director, RolEnum.secretaria):
        raise HTTPException(status_code=403, detail="No tiene permisos")


# ---------------------------------------------------------------------------
# Helpers for space (aula / laboratorio) and docente schedules
# ---------------------------------------------------------------------------

def _bloques_aula(db: Session, aula_id: int, semestre_id: int):
    bloques = (
        db.query(BloqueHorario)
        .options(joinedload(BloqueHorario.asignacion).joinedload(AsignacionCarga.curso))
        .filter(BloqueHorario.aula_id == aula_id, BloqueHorario.semestre_id == semestre_id)
        .order_by(BloqueHorario.dia, BloqueHorario.hora_inicio)
        .all()
    )
    ids = {b.docente_id for b in bloques if b.docente_id}
    dm = {d.id: d for d in db.query(Docente).filter(Docente.id.in_(ids)).all()} if ids else {}
    return bloques, dm


def _bloques_laboratorio_espacio(db: Session, laboratorio_id: int, semestre_id: int):
    bloques = (
        db.query(BloqueHorario)
        .options(joinedload(BloqueHorario.asignacion).joinedload(AsignacionCarga.curso))
        .filter(BloqueHorario.laboratorio_id == laboratorio_id, BloqueHorario.semestre_id == semestre_id)
        .order_by(BloqueHorario.dia, BloqueHorario.hora_inicio)
        .all()
    )
    ids = {b.docente_id for b in bloques if b.docente_id}
    dm = {d.id: d for d in db.query(Docente).filter(Docente.id.in_(ids)).all()} if ids else {}
    return bloques, dm


def _espacio_label(b, docentes_map) -> str:
    parts = []
    if b.asignacion:
        parts.append(b.asignacion.curso.codigo)
    d = docentes_map.get(b.docente_id)
    if d:
        parts.append(d.apellidos[:14])
    return "\n".join(parts) if parts else "---"


def _docente_label(b) -> str:
    parts = [TIPO_LABEL.get(b.tipo, b.tipo)]
    if b.asignacion:
        parts.append(b.asignacion.curso.codigo)
    if b.aula:
        parts.append(b.aula.nombre[:10])
    elif b.laboratorio:
        parts.append(b.laboratorio.nombre[:10])
    return "\n".join(parts)


def _build_weekly_grid_pdf(title: str, subtitle: str, bloques: list, label_fn) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Normal"], fontSize=13,
                                 alignment=TA_CENTER, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("s", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER)
    cell_style = ParagraphStyle("c", parent=styles["Normal"], fontSize=7)

    story = [
        Paragraph("UNIVERSIDAD NACIONAL DE TUCUMAN", title_style),
        Paragraph(title, title_style),
        Paragraph(subtitle, sub_style),
        Spacer(1, 0.4*cm),
    ]

    HORAS = list(range(7, 21))
    header = ["Hora"] + [DIAS_LABEL[d] for d in DIAS]
    grid_data = [header]
    for h in HORAS:
        row = [f"{h:02d}:00"]
        for dia in DIAS:
            here = [b for b in bloques if b.dia == dia
                    and _time_to_float(b.hora_inicio) <= h < _time_to_float(b.hora_fin)]
            row.append(Paragraph(label_fn(here[0]), cell_style) if here else "")
        grid_data.append(row)

    col_w = [1.5*cm] + [3.5*cm] * 6
    grid = Table(grid_data, colWidths=col_w, rowHeights=[0.55*cm] * (len(HORAS) + 1))
    ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F9FAFB")),
    ])
    for ri in range(1, len(grid_data)):
        for ci in range(1, 7):
            dia = DIAS[ci - 1]
            h = HORAS[ri - 1]
            here = [b for b in bloques if b.dia == dia
                    and _time_to_float(b.hora_inicio) <= h < _time_to_float(b.hora_fin)]
            if here:
                rgb = TIPO_COLOR.get(here[0].tipo, (1, 1, 1))
                ts.add("BACKGROUND", (ci, ri), (ci, ri), colors.Color(*rgb))
    grid.setStyle(ts)
    story.append(grid)
    doc.build(story)
    return buf.getvalue()


def _build_weekly_grid_excel(ws, title: str, subtitle: str, bloques: list, label_fn):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:H2")
    ws["A2"] = subtitle
    ws["A2"].alignment = Alignment(horizontal="center")

    grid_header_fill = PatternFill("solid", fgColor="374151")
    row = 4
    ws.cell(row=row, column=1, value="Hora").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = grid_header_fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    for ci, dia in enumerate(DIAS, 2):
        cell = ws.cell(row=row, column=ci, value=DIAS_LABEL[dia])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = grid_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    TIPO_FILL = {k: PatternFill("solid", fgColor="%02X%02X%02X" % (
                    int(v[0]*255), int(v[1]*255), int(v[2]*255)))
                 for k, v in TIPO_COLOR.items()}

    row += 1
    for h in range(7, 21):
        ws.cell(row=row, column=1, value=f"{h:02d}:00").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = border
        for ci, dia in enumerate(DIAS, 2):
            here = [b for b in bloques if b.dia == dia
                    and _time_to_float(b.hora_inicio) <= h < _time_to_float(b.hora_fin)]
            cell = ws.cell(row=row, column=ci)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if here:
                cell.value = label_fn(here[0])
                cell.fill = TIPO_FILL.get(here[0].tipo, PatternFill())
        row += 1

    ws.column_dimensions["A"].width = 10
    for ci in range(2, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 18


# ---------------------------------------------------------------------------
# Horario por aula
# ---------------------------------------------------------------------------

@router.get("/horario-por-aula/{semestre_id}/{aula_id}/pdf")
def horario_aula_pdf(
    semestre_id: int,
    aula_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_secretaria),
):
    semestre = _get_semestre(db, semestre_id)
    aula = db.get(Aula, aula_id)
    if not aula:
        raise HTTPException(status_code=404, detail="Aula no encontrada")
    bloques, dm = _bloques_aula(db, aula_id, semestre_id)
    pdf = _build_weekly_grid_pdf(
        f"Horario -- {aula.nombre}",
        f"Semestre {semestre.numero} {semestre.anio}",
        bloques,
        lambda b: _espacio_label(b, dm),
    )
    fname = f"horario_aula_{aula.nombre.replace(' ', '_')}.pdf"
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/horario-por-aula/{semestre_id}/{aula_id}/excel")
def horario_aula_excel(
    semestre_id: int,
    aula_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_secretaria),
):
    import openpyxl
    semestre = _get_semestre(db, semestre_id)
    aula = db.get(Aula, aula_id)
    if not aula:
        raise HTTPException(status_code=404, detail="Aula no encontrada")
    bloques, dm = _bloques_aula(db, aula_id, semestre_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = aula.nombre[:30]
    _build_weekly_grid_excel(
        ws,
        f"Horario -- {aula.nombre}",
        f"Semestre {semestre.numero} {semestre.anio}",
        bloques,
        lambda b: _espacio_label(b, dm),
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"horario_aula_{aula.nombre.replace(' ', '_')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------------------
# Horario por laboratorio
# ---------------------------------------------------------------------------

@router.get("/horario-por-laboratorio/{semestre_id}/{laboratorio_id}/pdf")
def horario_laboratorio_pdf(
    semestre_id: int,
    laboratorio_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_secretaria),
):
    semestre = _get_semestre(db, semestre_id)
    lab = db.get(Laboratorio, laboratorio_id)
    if not lab:
        raise HTTPException(status_code=404, detail="Laboratorio no encontrado")
    bloques, dm = _bloques_laboratorio_espacio(db, laboratorio_id, semestre_id)
    pdf = _build_weekly_grid_pdf(
        f"Horario -- {lab.nombre}",
        f"Semestre {semestre.numero} {semestre.anio}",
        bloques,
        lambda b: _espacio_label(b, dm),
    )
    fname = f"horario_lab_{lab.nombre.replace(' ', '_')}.pdf"
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/horario-por-laboratorio/{semestre_id}/{laboratorio_id}/excel")
def horario_laboratorio_excel(
    semestre_id: int,
    laboratorio_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_secretaria),
):
    import openpyxl
    semestre = _get_semestre(db, semestre_id)
    lab = db.get(Laboratorio, laboratorio_id)
    if not lab:
        raise HTTPException(status_code=404, detail="Laboratorio no encontrado")
    bloques, dm = _bloques_laboratorio_espacio(db, laboratorio_id, semestre_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = lab.nombre[:30]
    _build_weekly_grid_excel(
        ws,
        f"Horario -- {lab.nombre}",
        f"Semestre {semestre.numero} {semestre.anio}",
        bloques,
        lambda b: _espacio_label(b, dm),
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"horario_lab_{lab.nombre.replace(' ', '_')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------------------
# Mi horario (docente's own schedule)
# ---------------------------------------------------------------------------

@router.get("/mi-horario/{docente_id}/pdf")
def mi_horario_pdf(
    docente_id: int,
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_active_user),
):
    _check_docente_or_staff(current, docente_id, db)
    semestre = db.query(Semestre).filter(Semestre.activo == True).first()
    if not semestre:
        raise HTTPException(status_code=409, detail="No hay semestre activo")
    docente = _get_docente(db, docente_id)
    bloques = _bloques_docente(db, docente_id, semestre.id)
    pdf = _build_weekly_grid_pdf(
        f"Horario -- {docente.apellidos}, {docente.nombre}",
        f"Semestre {semestre.numero} {semestre.anio}",
        bloques,
        _docente_label,
    )
    fname = f"horario_{docente.apellidos}_{docente.nombre}.pdf".replace(" ", "_")
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/mi-horario/{docente_id}/excel")
def mi_horario_excel(
    docente_id: int,
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_active_user),
):
    import openpyxl
    _check_docente_or_staff(current, docente_id, db)
    semestre = db.query(Semestre).filter(Semestre.activo == True).first()
    if not semestre:
        raise HTTPException(status_code=409, detail="No hay semestre activo")
    docente = _get_docente(db, docente_id)
    bloques = _bloques_docente(db, docente_id, semestre.id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horario"
    _build_weekly_grid_excel(
        ws,
        f"Horario -- {docente.apellidos}, {docente.nombre}",
        f"Semestre {semestre.numero} {semestre.anio}",
        bloques,
        _docente_label,
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"horario_{docente.apellidos}_{docente.nombre}.xlsx".replace(" ", "_")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})

# ---------------------------------------------------------------------------
# Plan de Estudios
# ---------------------------------------------------------------------------

@router.get("/plan-estudios/pdf")
def plan_estudios_pdf(
    db: Session = Depends(get_db),
    current: Usuario = Depends(get_current_active_user),
):
    if current.rol not in (RolEnum.admin, RolEnum.director, RolEnum.secretaria):
        raise HTTPException(status_code=403, detail="No tiene permisos")
        
    cursos = (
        db.query(Curso)
        .options(joinedload(Curso.departamento), joinedload(Curso.escuela))
        .all()
    )
    
    if not cursos:
        raise HTTPException(status_code=404, detail="No hay cursos registrados")

    from pdf_plan_estudios import generate_plan_estudios_pdf
    
    # Asumimos una escuela para el título si todos son de la misma, o la primera que encontremos
    escuela_nombre = cursos[0].escuela.nombre if cursos and cursos[0].escuela else "SISTEMAS"
    
    pdf_bytes = generate_plan_estudios_pdf(cursos, escuela_nombre=escuela_nombre)
    fname = "plan_de_estudios.pdf"
    
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )
